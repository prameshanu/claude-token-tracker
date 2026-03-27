from __future__ import annotations

import argparse
import logging
import os
import threading
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, numbers

logger = logging.getLogger("claude_token_tracker")

HEADERS = [
    "Timestamp",
    "Request ID",
    "Model",
    "Input Tokens",
    "Output Tokens",
    "Total Tokens",
    "Cache Read Tokens",
    "Cache Creation Tokens",
    "Input Cost ($)",
    "Output Cost ($)",
    "Total Cost ($)",
    "Task Label",
    "Project",
    "Method",
    "Duration (ms)",
]

# Lock to prevent concurrent file writes from corrupting the Excel file
_file_lock = threading.Lock()


def _style_header(ws) -> None:
    """Apply styling to the header row."""
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")


def _get_or_create_workbook(path: str) -> tuple[Workbook, Any]:
    """Load existing workbook or create a new one with headers."""
    filepath = Path(path)
    if filepath.exists() and filepath.stat().st_size > 0:
        wb = load_workbook(filepath)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Token Usage"
        _style_header(ws)
        # Set column widths
        widths = [20, 20, 30, 14, 14, 14, 18, 20, 14, 14, 14, 20, 20, 10, 14]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    return wb, ws


def append_row(path: str, row: dict[str, Any]) -> None:
    """Append a single usage row to the Excel file (thread-safe)."""
    with _file_lock:
        wb, ws = _get_or_create_workbook(path)
        total_tokens = row.get("input_tokens", 0) + row.get("output_tokens", 0)
        total_cost = row.get("input_cost", 0) + row.get("output_cost", 0)
        values = [
            datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S"),
            row.get("request_id", ""),
            row.get("model", ""),
            row.get("input_tokens", 0),
            row.get("output_tokens", 0),
            total_tokens,
            row.get("cache_read_tokens", 0),
            row.get("cache_creation_tokens", 0),
            row.get("input_cost", 0),
            row.get("output_cost", 0),
            round(total_cost, 6),
            row.get("task_label", ""),
            row.get("project", ""),
            row.get("method", ""),
            row.get("duration_ms", 0),
        ]
        ws.append(values)

        # Format cost cells as currency
        last_row = ws.max_row
        for col in (9, 10, 11):  # input_cost, output_cost, total_cost
            ws.cell(row=last_row, column=col).number_format = numbers.FORMAT_NUMBER_00

        wb.save(path)


def append_row_background(path: str, row: dict[str, Any]) -> None:
    """Fire-and-forget Excel append in a daemon thread."""
    t = threading.Thread(target=_safe_append, args=(path, row), daemon=True)
    t.start()


def _safe_append(path: str, row: dict[str, Any]) -> None:
    try:
        append_row(path, row)
    except Exception:
        logger.exception("Failed to log token usage to Excel at %s", path)


def export_from_mysql(config=None, output_path: str | None = None) -> str:
    """Export all rows from MySQL to a new Excel file.

    Returns the path to the created file.
    """
    from claude_token_tracker.config import TrackerConfig
    from claude_token_tracker.db import TokenDB

    cfg = config or TrackerConfig.from_env()
    output = output_path or cfg.excel_path
    db = TokenDB(cfg)
    pool = db._get_pool()
    conn = pool.get_connection()

    try:
        cursor = conn.cursor(dictionary=True)
        cursor.execute(
            "SELECT request_id, model, input_tokens, output_tokens, "
            "cache_read_tokens, cache_creation_tokens, "
            "input_cost, output_cost, task_label, project, method, "
            "duration_ms, created_at FROM claude_token_usage ORDER BY created_at"
        )
        rows = cursor.fetchall()
    finally:
        conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Token Usage"
    _style_header(ws)

    widths = [20, 20, 30, 14, 14, 14, 18, 20, 14, 14, 14, 20, 20, 10, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    for row in rows:
        total_tokens = (row.get("input_tokens", 0) or 0) + (row.get("output_tokens", 0) or 0)
        total_cost = float(row.get("input_cost", 0) or 0) + float(row.get("output_cost", 0) or 0)
        created = row.get("created_at", "")
        if hasattr(created, "strftime"):
            created = created.strftime("%Y-%m-%d %H:%M:%S")
        values = [
            str(created),
            row.get("request_id", ""),
            row.get("model", ""),
            row.get("input_tokens", 0),
            row.get("output_tokens", 0),
            total_tokens,
            row.get("cache_read_tokens", 0),
            row.get("cache_creation_tokens", 0),
            float(row.get("input_cost", 0) or 0),
            float(row.get("output_cost", 0) or 0),
            round(total_cost, 6),
            row.get("task_label", ""),
            row.get("project", ""),
            row.get("method", ""),
            row.get("duration_ms", 0),
        ]
        ws.append(values)

    # Format cost columns
    for row_idx in range(2, ws.max_row + 1):
        for col in (9, 10, 11):
            ws.cell(row=row_idx, column=col).number_format = numbers.FORMAT_NUMBER_00

    wb.save(output)
    return os.path.abspath(output)


def cli_export() -> None:
    """CLI entry point: claude-tracker-export"""
    parser = argparse.ArgumentParser(description="Export Claude token usage from MySQL to Excel")
    parser.add_argument("-o", "--output", default=None, help="Output Excel file path")
    args = parser.parse_args()
    path = export_from_mysql(output_path=args.output)
    print(f"Exported to: {path}")
